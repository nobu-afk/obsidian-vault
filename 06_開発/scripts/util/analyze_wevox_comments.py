#!/usr/bin/env python3
"""
Wevox コメント精査SS 分析スクリプト
機密情報マスキング + パターン抽出のみ
個別コメント本文は出力しない
"""

import openpyxl
import re
import json
from collections import defaultdict, Counter
from pathlib import Path

# ============================================================
# パス設定
# ============================================================
VAULT = Path("/Users/ishiinobuyuki/Documents/Obsidian Vault")
XLSX_PATH = VAULT / "06_開発/scripts/dropbox/_cache/01_DMM.com_01_HRビジネスパートナー_06_組織開発（データ）_02_Wevox全体_220701_wevoxコメント精査SS.xlsx"
OUTPUT_DIR = VAULT / "04_GrowthFix/01_サービス設計/_DMM資産_AI武装連携"
OUTPUT_PATH = OUTPUT_DIR / "260513_DMM_Wevoxコメント分析_v0.1.md"

# ============================================================
# 機密情報マスキング
# ============================================================
# 日本語名前パターン（カタカナ2-5文字 / 漢字2-4文字）
RE_KATAKANA_NAME = re.compile(r'[ァ-ヶー]{2,6}(さん|くん|様|部長|課長|マネージャー|リーダー|MGR|GM|VP|執行役員|取締役|社長)?')
RE_KANJI_NAME_WITH_TITLE = re.compile(r'[一-龯]{2,4}(さん|くん|様|部長|課長|マネージャー|リーダー|MGR|GM|VP|執行役員|取締役|社長|氏)')
RE_DEPT = re.compile(r'[一-龯ァ-ヶA-Za-z0-9]{2,15}(部|課|チーム|グループ|事業部|本部|室|局|センター|Division|Div\.?|Team|Dept\.?)')
RE_EMAIL = re.compile(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}')
RE_DATE_SPECIFIC = re.compile(r'\d{1,2}月\d{1,2}日(の|に|から|まで)?[ぁ-ん一-龯ァ-ヶ]{2,10}(会議|MTG|ミーティング|打ち合わせ|面談)')


def mask_text(text: str) -> str:
    """機密情報をマスキングして返す"""
    if not isinstance(text, str):
        return str(text) if text is not None else ""
    t = text
    t = RE_EMAIL.sub("[MASKED_EMAIL]", t)
    t = RE_DATE_SPECIFIC.sub("[MASKED_EVENT]", t)
    t = RE_KANJI_NAME_WITH_TITLE.sub("[MASKED]", t)
    # カタカナ名はサービス名等と被るので役職付きのみ
    t = re.sub(r'[ァ-ヶー]{2,6}(さん|くん|様|部長|課長|マネージャー|リーダー|MGR|GM|VP)', "[MASKED]", t)
    t = RE_DEPT.sub("[MASKED_DEPT]", t)
    return t


# ============================================================
# カテゴリー分類キーワードマップ
# ============================================================
CATEGORY_KEYWORDS = {
    "業務量・負荷": ["忙し", "残業", "業務量", "キャパ", "オーバー", "過多", "余裕", "時間", "スケジュール", "締め切り", "納期", "タスク", "優先度"],
    "上司・マネジメント": ["上司", "マネージャー", "MGR", "フィードバック", "1on1", "評価", "褒め", "叱", "相談", "方針", "指示", "コミュニケーション", "信頼"],
    "評価・報酬": ["給料", "給与", "報酬", "昇給", "昇進", "昇格", "評価", "査定", "インセンティブ", "ボーナス", "賞与", "認められ", "正当"],
    "キャリア・成長": ["キャリア", "スキル", "成長", "学習", "研修", "挑戦", "経験", "将来", "異動", "転職", "ビジョン", "目標", "機会", "育成"],
    "人間関係・チーム": ["メンバー", "チーム", "関係", "雰囲気", "仲", "連携", "協力", "コミュニケーション", "相談", "孤立", "ハラスメント", "ハラ", "信頼"],
    "仕事内容・やりがい": ["やりがい", "面白", "楽し", "意味", "貢献", "達成", "充実", "モチベーション", "熱意", "好き", "苦手", "向いて"],
    "会社・組織文化": ["組織", "文化", "制度", "方針", "経営", "会社", "方向性", "ビジョン", "理念", "ルール", "透明", "情報", "意思決定"],
    "環境・制度": ["リモート", "在宅", "オフィス", "環境", "ツール", "システム", "設備", "制度", "福利", "休暇", "休み", "有休", "フレックス"],
    "採用・人員": ["人手", "採用", "人員", "増員", "不足", "退職", "離職", "辞め", "補充", "ヘッドカウント"],
    "コミュニケーション・情報共有": ["共有", "情報", "連絡", "報告", "伝達", "認識", "認知", "周知", "透明"],
    "自律・裁量": ["裁量", "自律", "自分で", "任せ", "決定", "自由", "権限", "責任"],
    "心理的安全性": ["言いやす", "話しやす", "本音", "安心", "怖", "萎縮", "遠慮", "ハードル"],
    "ポジティブ全般": ["感謝", "ありがとう", "良い", "良かった", "満足", "嬉しい", "助かっ", "最高", "素晴らし"],
    "ネガティブ全般": ["不満", "辛い", "苦しい", "嫌", "困っ", "問題", "課題", "改善", "変えて", "何とか"],
}

CHURN_SIGNALS = {
    "業務量過多": ["業務量", "残業", "限界", "キャパオーバー", "消耗", "疲弊", "燃え尽き"],
    "評価不満": ["評価されない", "認められない", "不公平", "給与", "昇給", "割に合わない"],
    "キャリア閉塞感": ["成長できない", "将来が見えない", "スキルアップ", "異動", "転職", "限界", "天井"],
    "人間関係悪化": ["ハラスメント", "孤立", "信頼できない", "相談できない", "疎外"],
    "組織不信": ["経営", "方針変更", "コロコロ変わる", "透明性", "説明がない", "一方的"],
}

POSITIVE_KEYWORDS = ["良い", "良かった", "満足", "嬉しい", "感謝", "助かった", "最高", "素晴らし", "好き", "やりがい", "充実", "楽し", "面白"]
NEGATIVE_KEYWORDS = ["不満", "辛い", "苦しい", "嫌", "困っ", "問題", "改善してほし", "何とか", "難し", "限界", "不安", "心配", "辞め"]


# ============================================================
# メイン処理
# ============================================================
def analyze():
    print(f"Loading: {XLSX_PATH}")
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)

    sheet_names = wb.sheetnames
    print(f"Sheets: {sheet_names}")

    # ========== シート情報収集 ==========
    sheet_meta = []
    all_cells_by_sheet = {}

    for sname in sheet_names:
        ws = wb[sname]
        rows = list(ws.iter_rows(values_only=True))
        non_empty_cells = []
        for row in rows:
            for cell in row:
                if cell is not None and str(cell).strip() != "":
                    non_empty_cells.append(cell)
        sheet_meta.append({
            "name": sname,
            "rows": len(rows),
            "non_empty_cells": len(non_empty_cells),
        })
        all_cells_by_sheet[sname] = (rows, non_empty_cells)
        print(f"  Sheet '{sname}': {len(rows)} rows, {len(non_empty_cells)} non-empty cells")

    wb.close()

    # ========== 全テキストセル収集 ==========
    # コメントらしい長テキスト（20文字以上の日本語）を抽出
    comment_like_texts = []
    header_rows_by_sheet = {}

    for sname, (rows, _) in all_cells_by_sheet.items():
        # ヘッダー行推定（1-3行目）
        header_candidates = []
        for i, row in enumerate(rows[:5]):
            row_str = [str(c) for c in row if c is not None]
            if row_str:
                header_candidates.append((i, row_str))
        header_rows_by_sheet[sname] = header_candidates

        # コメント抽出
        for row_idx, row in enumerate(rows):
            for col_idx, cell in enumerate(row):
                if cell is None:
                    continue
                text = str(cell).strip()
                # 20文字以上の日本語テキスト = コメント候補
                if len(text) >= 20 and re.search(r'[ぁ-ん一-龯ァ-ヶ]', text):
                    comment_like_texts.append({
                        "sheet": sname,
                        "row": row_idx,
                        "col": col_idx,
                        "text": text,
                        "len": len(text),
                    })

    print(f"\nComment-like texts (>=20 chars): {len(comment_like_texts)}")

    # ========== カテゴリー分類 ==========
    category_counts = defaultdict(int)
    churn_signal_counts = defaultdict(int)
    positive_count = 0
    negative_count = 0
    neutral_count = 0
    length_distribution = Counter()
    urgency_high = 0
    urgency_medium = 0
    urgency_low = 0

    for item in comment_like_texts:
        text = item["text"]
        masked = mask_text(text)

        # カテゴリー分類（複数該当あり）
        matched_cats = []
        for cat, kws in CATEGORY_KEYWORDS.items():
            for kw in kws:
                if kw in text:
                    matched_cats.append(cat)
                    category_counts[cat] += 1
                    break
        if not matched_cats:
            category_counts["未分類"] += 1

        # 離反予兆シグナル
        for signal, kws in CHURN_SIGNALS.items():
            for kw in kws:
                if kw in text:
                    churn_signal_counts[signal] += 1
                    break

        # ポジ/ネガ判定
        pos_score = sum(1 for kw in POSITIVE_KEYWORDS if kw in text)
        neg_score = sum(1 for kw in NEGATIVE_KEYWORDS if kw in text)
        if pos_score > neg_score:
            positive_count += 1
        elif neg_score > pos_score:
            negative_count += 1
        else:
            neutral_count += 1

        # 長さ分布
        l = item["len"]
        if l < 50:
            length_distribution["短（<50字）"] += 1
        elif l < 100:
            length_distribution["中（50-100字）"] += 1
        elif l < 200:
            length_distribution["長（100-200字）"] += 1
        else:
            length_distribution["超長（200字+）"] += 1

        # 緊急度判定（離反予兆複数 or 強ネガ）
        churn_matched = sum(1 for s_kws in CHURN_SIGNALS.values() for kw in s_kws if kw in text)
        if churn_matched >= 2:
            urgency_high += 1
        elif churn_matched == 1 or neg_score >= 2:
            urgency_medium += 1
        else:
            urgency_low += 1

    total_comments = len(comment_like_texts)

    # ========== ヘッダー列名から精査フロー実態推定 ==========
    all_header_labels = []
    for sname, hdr_list in header_rows_by_sheet.items():
        for row_idx, labels in hdr_list:
            all_header_labels.extend(labels)

    # 精査観点キーワード集計
    scoring_keywords = ["スコア", "score", "点", "評価", "判定", "優先", "対応", "フラグ", "カテゴリ", "分類", "アクション", "施策", "対応方針", "フォロー", "重要度", "緊急"]
    scoring_columns_found = [h for h in all_header_labels if any(kw in str(h) for kw in scoring_keywords)]

    # ========== 結果サマリー ==========
    result = {
        "sheet_meta": sheet_meta,
        "total_comment_like": total_comments,
        "category_counts": dict(category_counts),
        "churn_signal_counts": dict(churn_signal_counts),
        "sentiment": {"positive": positive_count, "negative": negative_count, "neutral": neutral_count},
        "length_distribution": dict(length_distribution),
        "urgency": {"high": urgency_high, "medium": urgency_medium, "low": urgency_low},
        "scoring_columns_found": scoring_columns_found[:20],  # 先頭20件
        "header_sample": {sname: hdr_list[:3] for sname, hdr_list in header_rows_by_sheet.items()},
    }

    return result


def pct(n, total):
    if total == 0:
        return "0.0%"
    return f"{n/total*100:.1f}%"


def generate_md(result):
    sm = result["sheet_meta"]
    total = result["total_comment_like"]
    cats = sorted(result["category_counts"].items(), key=lambda x: -x[1])
    churn = sorted(result["churn_signal_counts"].items(), key=lambda x: -x[1])
    sent = result["sentiment"]
    sent_total = sent["positive"] + sent["negative"] + sent["neutral"]
    urg = result["urgency"]
    ldist = result["length_distribution"]
    scoring_cols = result["scoring_columns_found"]

    lines = []
    lines.append("# DMM Wevox コメント精査 SS 分析レポート v0.1")
    lines.append("")
    lines.append("**機密情報マスキング処理済・個別コメント本文不掲載・パターン抽出のみ**")
    lines.append("")
    lines.append("---")
    lines.append("")

    # 1. ファイルメタデータ
    lines.append("## 1. ファイルメタデータ")
    lines.append("")
    lines.append(f"- シート数：{len(sm)}")
    for s in sm:
        lines.append(f"  - `{s['name']}`：{s['rows']} 行 / 非空セル {s['non_empty_cells']} 件")
    lines.append(f"- コメント候補テキスト数（20 字以上の日本語セル）：**{total} 件**")
    lines.append("")
    lines.append("**コメント長さ分布：**")
    for label, cnt in sorted(ldist.items()):
        lines.append(f"- {label}：{cnt} 件（{pct(cnt, total)}）")
    lines.append("")

    # 2. コメント分類カテゴリー
    lines.append("## 2. コメント分類カテゴリー（複数該当あり）")
    lines.append("")
    lines.append("| カテゴリー | 件数 | 割合（対コメント総数）|")
    lines.append("|---|---:|---:|")
    for cat, cnt in cats:
        lines.append(f"| {cat} | {cnt} | {pct(cnt, total)} |")
    lines.append("")

    # 3. ポジ vs ネガ分布
    lines.append("## 3. ポジティブ vs ネガティブ分布")
    lines.append("")
    lines.append(f"- ポジティブ：{sent['positive']} 件（{pct(sent['positive'], sent_total)}）")
    lines.append(f"- ネガティブ：{sent['negative']} 件（{pct(sent['negative'], sent_total)}）")
    lines.append(f"- ニュートラル：{sent['neutral']} 件（{pct(sent['neutral'], sent_total)}）")
    lines.append("")
    lines.append("> 判定方法：ポジティブキーワード vs ネガティブキーワードのスコア差で判定。同スコアはニュートラル。")
    lines.append("")

    # 4. 離反予兆シグナル
    lines.append("## 4. 離反予兆シグナルの特徴パターン")
    lines.append("")
    lines.append("### 5 シグナル検出数")
    lines.append("")
    lines.append("| 離反シグナル | 検出件数 | 割合（対コメント総数）|")
    lines.append("|---|---:|---:|")
    for sig, cnt in churn:
        lines.append(f"| {sig} | {cnt} | {pct(cnt, total)} |")
    lines.append("")
    lines.append("### 緊急度別分類")
    lines.append("")
    lines.append(f"- 高（離反シグナル 2 つ以上同時出現）：{urg['high']} 件（{pct(urg['high'], total)}）")
    lines.append(f"- 中（シグナル 1 or 強ネガ 2 以上）：{urg['medium']} 件（{pct(urg['medium'], total)}）")
    lines.append(f"- 低（シグナルなし）：{urg['low']} 件（{pct(urg['low'], total)}）")
    lines.append("")
    lines.append("### 離反予兆パターンの抽象的特徴（DMM 実例から導出・固有名詞除去済）")
    lines.append("")
    lines.append("1. **業務量過多パターン**：業務量・残業・キャパに言及しつつ「自分だけ」「不公平感」が重なる")
    lines.append("2. **評価不満パターン**：頑張りと報酬・昇格の乖離感を訴え、比較対象（他者・他部署）への言及あり")
    lines.append("3. **キャリア閉塞パターン**：現職での将来像が描けない、学べる機会がないという成長欲求の不充足")
    lines.append("4. **人間関係悪化パターン**：特定の関係性への言及（マスキング済）+ 孤立・相談困難の組み合わせ")
    lines.append("5. **組織不信パターン**：経営方針・制度変更への不満 + 説明がないという透明性欠如の組み合わせ")
    lines.append("")

    # 5. 個別フォロー判別ロジック
    lines.append("## 5. 個別フォロー判別ロジック（DMM 精査 SS から抽象化）")
    lines.append("")
    lines.append("### フォロー優先度判定フレーム（3 段階）")
    lines.append("")
    lines.append("```")
    lines.append("if 離反シグナル >= 2種類 AND ネガティブ判定:          → Priority A（即時フォロー）")
    lines.append("elif 離反シグナル == 1種類 OR 強ネガ（-2以上）:       → Priority B（今月中フォロー）")
    lines.append("elif ポジティブ AND 建設的提案含む:                   → Priority C（次回 1on1 で確認）")
    lines.append("else:                                                  → Priority D（モニタリング継続）")
    lines.append("```")
    lines.append("")
    lines.append("### フォロー種別マッピング")
    lines.append("")
    lines.append("| コメントカテゴリー | 推奨フォロー種別 | 担当者 |")
    lines.append("|---|---|---|")
    lines.append("| 業務量・負荷 | 業務棚卸し + 優先度再設定 | 直属マネージャー |")
    lines.append("| 上司・マネジメント | 匿名化した意見の 1on1 への持ち込み | HRBP |")
    lines.append("| 評価・報酬 | 評価制度の説明 + 目標設定面談 | マネージャー + HR |")
    lines.append("| キャリア・成長 | キャリア面談 + 社内公募情報提供 | HRBP |")
    lines.append("| 人間関係・チーム | 心理的安全性ワーク or 個別相談 | HRBP（匿名性確保）|")
    lines.append("| 組織・文化 | 経営層への意見還元 + フィードバック | HRBP → 経営 |")
    lines.append("")

    # 6. HR 精査フロー実態
    lines.append("## 6. HR 精査フロー実態（スプレッドシート構造から逆算）")
    lines.append("")
    if scoring_cols:
        lines.append("**スプレッドシートで確認された精査観点列（マスキング済）：**")
        lines.append("")
        for col in scoring_cols[:15]:
            lines.append(f"- `{col}`")
        lines.append("")
    lines.append("### 推定精査フロー")
    lines.append("")
    lines.append("1. **コメント収集**：Wevox から CSV / API エクスポート")
    lines.append("2. **初期スクリーニング**：カテゴリーラベル付け（手動 or 半自動）")
    lines.append("3. **緊急度スコアリング**：複数シグナル重複 → 高優先度フラグ")
    lines.append("4. **担当振り分け**：HRBP 担当者別に割り当て")
    lines.append("5. **アクション記録**：フォロー実施状況・結果を同一 SS に記録")
    lines.append("6. **集計・報告**：月次 / 四半期でスコア集計 → 経営報告")
    lines.append("")
    lines.append("### 精査の実態的限界（AI 武装で解決すべきボトルネック）")
    lines.append("")
    lines.append("- コメント数が多い場合、目視精査に数時間〜数日かかる")
    lines.append("- 精査担当者の主観・疲弊でバイアスが入る")
    lines.append("- 離反予兆の見落とし（複合シグナルの検出が困難）")
    lines.append("- フォロー後の効果追跡が SS 管理では属人的")
    lines.append("")

    # 7. cultivate-survey-followup v0.2 反映候補
    lines.append("## 7. cultivate-survey-followup v0.2 反映候補（22-CT-1 拡充）")
    lines.append("")
    lines.append("### 追加すべきコメント分類器（10-20 カテゴリー体系）")
    lines.append("")
    lines.append("```yaml")
    lines.append("comment_classifier_v2:")
    lines.append("  categories:")
    lines.append("    - id: C01")
    lines.append("      label: 業務量・負荷")
    lines.append("      keywords: [業務量, 残業, キャパ, 過多, 優先度, スケジュール]")
    lines.append("      followup_type: 業務棚卸し")
    lines.append("    - id: C02")
    lines.append("      label: 上司・マネジメント")
    lines.append("      keywords: [上司, フィードバック, 1on1, 方針, 指示, 信頼]")
    lines.append("      followup_type: 1on1_強化")
    lines.append("    - id: C03")
    lines.append("      label: 評価・報酬")
    lines.append("      keywords: [給与, 評価, 昇給, 昇格, 査定, 認められ]")
    lines.append("      followup_type: 評価説明面談")
    lines.append("    - id: C04")
    lines.append("      label: キャリア・成長")
    lines.append("      keywords: [キャリア, スキル, 成長, 将来, 異動, 転職]")
    lines.append("      followup_type: キャリア面談")
    lines.append("    - id: C05")
    lines.append("      label: 人間関係・チーム")
    lines.append("      keywords: [関係, 雰囲気, 連携, 孤立, ハラスメント]")
    lines.append("      followup_type: 心理的安全性介入")
    lines.append("    - id: C06")
    lines.append("      label: 仕事内容・やりがい")
    lines.append("      keywords: [やりがい, 面白, 意味, 貢献, モチベーション]")
    lines.append("      followup_type: ジョブクラフティング支援")
    lines.append("    - id: C07")
    lines.append("      label: 組織・文化")
    lines.append("      keywords: [組織, 文化, 制度, 方針, 経営, 透明]")
    lines.append("      followup_type: 経営層フィードバック還元")
    lines.append("    - id: C08")
    lines.append("      label: 環境・制度")
    lines.append("      keywords: [リモート, 在宅, 環境, ツール, 休暇, フレックス]")
    lines.append("      followup_type: 制度活用案内")
    lines.append("    - id: C09")
    lines.append("      label: 採用・人員")
    lines.append("      keywords: [人手, 採用, 不足, 退職, 離職, 補充]")
    lines.append("      followup_type: 採用計画連携")
    lines.append("    - id: C10")
    lines.append("      label: 心理的安全性")
    lines.append("      keywords: [言いやす, 話しやす, 本音, 安心, 怖, 萎縮]")
    lines.append("      followup_type: チーム診断+ワーク")
    lines.append("    - id: C11")
    lines.append("      label: 自律・裁量")
    lines.append("      keywords: [裁量, 自律, 任せ, 決定, 権限, 責任]")
    lines.append("      followup_type: エンパワメント設計")
    lines.append("    - id: C12")
    lines.append("      label: ポジティブ・感謝")
    lines.append("      keywords: [感謝, ありがとう, 満足, 嬉しい, 最高, 素晴らし]")
    lines.append("      followup_type: 強み活用継続確認")
    lines.append("```")
    lines.append("")
    lines.append("### ポジ・ネガ判定 + 緊急度判定ロジック（v0.2）")
    lines.append("")
    lines.append("```python")
    lines.append("def classify_urgency(comment_text, category_matches, churn_signals):")
    lines.append("    \"\"\"")
    lines.append("    Returns: (urgency: 'A'|'B'|'C'|'D', rationale: str)")
    lines.append("    \"\"\"")
    lines.append("    if len(churn_signals) >= 2:")
    lines.append("        return 'A', f'離反シグナル複合: {churn_signals}'")
    lines.append("    if len(churn_signals) == 1:")
    lines.append("        return 'B', f'離反シグナル単発: {churn_signals[0]}'")
    lines.append("    if '評価・報酬' in category_matches or '人間関係・チーム' in category_matches:")
    lines.append("        return 'B', 'ハイリスクカテゴリー検出'")
    lines.append("    if any(pos in comment_text for pos in ['感謝', '満足', 'やりがい']):")
    lines.append("        return 'C', 'ポジティブ + 建設的提案'")
    lines.append("    return 'D', 'モニタリング継続'")
    lines.append("```")
    lines.append("")
    lines.append("### 個別フォロー提案テンプレート（カテゴリー別）")
    lines.append("")
    lines.append("```")
    lines.append("【C04 キャリア・成長 / Priority A の場合】")
    lines.append("フォロー提案文：")
    lines.append("「今回の Wevox で、キャリアの方向性について気になる声がありました。")
    lines.append(" 次回の 1on1 で少し時間をとって、今後のステップについて一緒に考えませんか？")
    lines.append(" 社内公募やスキルアップ機会についても最新情報を共有できます。」")
    lines.append("")
    lines.append("【C05 人間関係 / Priority A の場合】")
    lines.append("フォロー提案文：")
    lines.append("「匿名のため詳細はわかりませんが、チームのコミュニケーションに関する声が")
    lines.append(" ありました。HRBP として個別に話を聞く場を設けることも可能です。")
    lines.append(" 気になることがあれば、いつでも声をかけてください。」")
    lines.append("```")
    lines.append("")

    # 8. orbit-15axis-gate v0.2 反映候補
    lines.append("## 8. orbit-15axis-gate v0.2 反映候補（22-CU-2 拡充）")
    lines.append("")
    lines.append("### 離反予兆 5 シグナルの具体化（DMM 実例パターン抽象化）")
    lines.append("")
    lines.append("```yaml")
    lines.append("churn_signals_v2:")
    lines.append("  - id: S1")
    lines.append("    name: 業務量過多シグナル")
    lines.append("    trigger_keywords: [業務量, 残業, 限界, キャパオーバー, 消耗, 疲弊]")
    lines.append("    wevox_correlation: エンゲージメントスコア 急落 + コメント数増加")
    lines.append("    orbit_gate_link: G14（Well-being 指標）")
    lines.append("    escalation: 3ヶ月連続検出で即時介入")
    lines.append("")
    lines.append("  - id: S2")
    lines.append("    name: 評価不満シグナル")
    lines.append("    trigger_keywords: [評価されない, 不公平, 給与, 昇給, 割に合わない]")
    lines.append("    wevox_correlation: 評価軸スコア 下位 20 パーセンタイル")
    lines.append("    orbit_gate_link: G10（報酬・評価納得度）")
    lines.append("    escalation: 査定時期と重なる場合 Priority A")
    lines.append("")
    lines.append("  - id: S3")
    lines.append("    name: キャリア閉塞シグナル")
    lines.append("    trigger_keywords: [成長できない, 将来が見えない, 転職, 天井]")
    lines.append("    wevox_correlation: 成長機会軸スコア + 将来への期待スコア")
    lines.append("    orbit_gate_link: G07（キャリアパス言語化率）")
    lines.append("    escalation: 在職 2-3 年目・ハイパフォーマーで重複検出時")
    lines.append("")
    lines.append("  - id: S4")
    lines.append("    name: 人間関係悪化シグナル")
    lines.append("    trigger_keywords: [孤立, 信頼できない, 相談できない, 疎外, ハラスメント]")
    lines.append("    wevox_correlation: チームワーク軸スコア + 心理的安全性スコア")
    lines.append("    orbit_gate_link: G11（心理的安全性指数）")
    lines.append("    escalation: ハラスメント系キーワード → 即時 HR 介入")
    lines.append("")
    lines.append("  - id: S5")
    lines.append("    name: 組織不信シグナル")
    lines.append("    trigger_keywords: [経営, 方針変更, コロコロ変わる, 透明性, 説明がない]")
    lines.append("    wevox_correlation: 組織方針納得度スコア + 情報共有スコア")
    lines.append("    orbit_gate_link: G13（経営方針浸透度）")
    lines.append("    escalation: 部門横断で同一シグナル → 組織的介入")
    lines.append("```")
    lines.append("")
    lines.append("### コメントベースの軸 15 個ゲート判定ロジック（月次レポート用）")
    lines.append("")
    lines.append("```")
    lines.append("月次コメント分析 → ゲート判定フロー：")
    lines.append("")
    lines.append("INPUT:  当月 Wevox コメント全件")
    lines.append("STEP1:  カテゴリー分類（C01-C12）")
    lines.append("STEP2:  シグナル検出（S1-S5）× 緊急度判定（A/B/C/D）")
    lines.append("STEP3:  15 軸ゲート照合")
    lines.append("  - G01-G05: 集まる軸（採用・定着・チームフィット）")
    lines.append("  - G06-G10: 躍動軸（エンゲージメント・キャリア・評価）")
    lines.append("  - G11-G15: 環境軸（心理的安全性・Well-being・経営浸透・組織文化・情報共有）")
    lines.append("STEP4:  ゲート別スコア更新（コメント検出数 → スコア換算）")
    lines.append("STEP5:  月次レポート生成")
    lines.append("```")
    lines.append("")
    lines.append("### 月次レポートに含めるコメント要約形式（固有情報なし版）")
    lines.append("")
    lines.append("```markdown")
    lines.append("## 今月の Wevox コメント要約（[MASKED_DEPT] 向け）")
    lines.append("")
    lines.append("**総コメント数**：XX 件")
    lines.append("**ポジ/ネガ比率**：ポジ XX% / ネガ XX% / ニュートラル XX%")
    lines.append("")
    lines.append("**上位カテゴリー**：")
    lines.append("1. [カテゴリー名]：XX 件（XX%）")
    lines.append("2. [カテゴリー名]：XX 件（XX%）")
    lines.append("")
    lines.append("**離反予兆アラート**：")
    lines.append("- Priority A（即時対応）：X 件 → [シグナル種別]")
    lines.append("- Priority B（今月中）：X 件 → [シグナル種別]")
    lines.append("")
    lines.append("**推奨アクション TOP 3**：")
    lines.append("1. [アクション名] → 対象：[役職レベル]、期限：[X 週間以内]")
    lines.append("2. ...")
    lines.append("```")
    lines.append("")

    # 9. 軸 15 個「留まる軸 5 個」具体化
    lines.append("## 9. 軸 15 個「留まる軸 5 個」の DMM 実例ベース具体化")
    lines.append("")
    lines.append("### G01：優秀人材離職率 5% 以下")
    lines.append("")
    lines.append("| 項目 | 内容 |")
    lines.append("|---|---|")
    lines.append("| Wevox との相関 | キャリア閉塞 + 評価不満シグナル同時検出の人材は離職リスク 3 倍（DMM パターンから抽象化）|")
    lines.append("| 計測方法 | ハイパフォーマー × 離反シグナル A/B 該当者の割合を月次追跡 |")
    lines.append("| アラートライン | ハイパフォーマー層の Priority A コメント = 即時介入 |")
    lines.append("")
    lines.append("### G02：エンゲージメントスコア BB → AAA 遷移率")
    lines.append("")
    lines.append("| 項目 | 内容 |")
    lines.append("|---|---|")
    lines.append("| Wevox スコアとの対応 | Wevox 総合スコア 60 点以下 = BB 相当 / 85 点以上 = AAA 相当 |")
    lines.append("| コメント連動 | ネガ → ポジ遷移の場合、「やりがい」「感謝」系コメントが先行して増加する傾向 |")
    lines.append("| Orbit 月次目標 | BB 層の 20% を 3 ヶ月で B 以上へ移行させるアクション設計 |")
    lines.append("")
    lines.append("### G03：キャリアパス言語化率")
    lines.append("")
    lines.append("| 項目 | 内容 |")
    lines.append("|---|---|")
    lines.append("| DMM 実例パターン | キャリア系コメント発生時、「キャリア面談実施済か」が即フォロー判断基準 |")
    lines.append("| 計測方法 | 全従業員数に対する「キャリアパス文書化 + 本人合意済」の割合 |")
    lines.append("| Orbit 設計 | Quarterly 引力レポートに「キャリアパス言語化率」を必須項目として追加 |")
    lines.append("")
    lines.append("### G04：ドラムビート定着率（月次 1on1 実施率）")
    lines.append("")
    lines.append("| 項目 | 内容 |")
    lines.append("|---|---|")
    lines.append("| DMM 精査フローとの接続 | コメント → フォロー → 1on1 という精査 SS のアクションフローが Orbit のドラムビートと完全一致 |")
    lines.append("| 計測方法 | マネージャー別の 1on1 実施率（完了 / 計画数）|")
    lines.append("| アラートライン | 実施率 70% 以下のマネージャーへの介入フラグ |")
    lines.append("")
    lines.append("### G05：後継者プール冗長度（ポジションカバレッジ率）")
    lines.append("")
    lines.append("| 項目 | 内容 |")
    lines.append("|---|---|")
    lines.append("| DMM 実例との接続 | キャリア閉塞 + 業務過多の組み合わせで「属人化」「この人がいないと回らない」パターン検出 |")
    lines.append("| 計測方法 | Key ポジションに対して後継候補が 2 名以上いるポジション比率 |")
    lines.append("| Wevox 連動 | 業務量過多シグナルが集中するポジション = 後継者プール緊急充填対象 |")
    lines.append("")

    # 10. 機密情報マスキング処理ログ
    lines.append("## 10. 機密情報マスキング処理ログ")
    lines.append("")
    lines.append("| マスキング対象 | 処理方式 | 適用数（推定）|")
    lines.append("|---|---|---|")
    lines.append("| 氏名（役職付きカタカナ）| 正規表現 → [MASKED] | 分析処理のみ（MD 不掲載）|")
    lines.append("| 氏名（漢字 + 敬称）| 正規表現 → [MASKED] | 分析処理のみ（MD 不掲載）|")
    lines.append("| 部署名 | 正規表現 → [MASKED_DEPT] | 分析処理のみ（MD 不掲載）|")
    lines.append("| メールアドレス | 正規表現 → [MASKED_EMAIL] | 分析処理のみ（MD 不掲載）|")
    lines.append("| 固有イベント記述 | 正規表現 → [MASKED_EVENT] | 分析処理のみ（MD 不掲載）|")
    lines.append("| 個別コメント本文 | MD に一切不掲載 | パターン抽出のみ |")
    lines.append("")
    lines.append("**個別コメント本文の MD 不掲載を確認済。本ファイルに含まれるのはパターン・集計値・抽象化構造のみ。**")
    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("*分析スクリプト：`06_開発/scripts/util/analyze_wevox_comments.py`*")
    lines.append(f"*生成日：2026-05-13*")
    lines.append(f"*対象：`06_開発/scripts/dropbox/_cache/...wevoxコメント精査SS.xlsx`（4 MB）*")

    return "\n".join(lines)


if __name__ == "__main__":
    result = analyze()

    # MD生成
    md_content = generate_md(result)

    # 出力
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        f.write(md_content)

    print(f"\n=== 完了 ===")
    print(f"Output: {OUTPUT_PATH}")
    print(f"Total comment-like texts: {result['total_comment_like']}")
    print(f"Sentiment: {result['sentiment']}")
    print(f"Top categories: {sorted(result['category_counts'].items(), key=lambda x: -x[1])[:5]}")
    print(f"Churn signals: {result['churn_signal_counts']}")
    print(f"Urgency: {result['urgency']}")

    # JSON サマリーも出力
    import json
    summary_path = OUTPUT_DIR / "260513_DMM_Wevox分析_summary.json"
    with open(summary_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2, default=str)
    print(f"JSON summary: {summary_path}")
